"""Excel 读写（openpyxl）：下载文件的内容读取与写回，APS/WMS 导出断言的高频场景。

- read_xlsx: 读取 xlsx 首表（或指定表），首行作表头，返回结构化行数据（限行防撑爆上下文）。
- write_xlsx: 将二维列表写入 xlsx（首行可作表头），用于下载模板填数后再上传的写操作场景。
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


def read_xlsx(
    path: str | Path,
    sheet: str | None = None,
    max_rows: int = 200,
) -> dict:
    """读取 xlsx 内容，返回 {sheet, headers, rows, row_count, truncated}。

    sheet: 工作表名（None = 第一个工作表）。max_rows 限制返回数据行数，
    超出则 truncated=True（不含表头行），避免大表撑爆上下文。
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.worksheets[0]
        headers: list[str] = []
        rows: list[list[Any]] = []
        truncated = False
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = ["" if c is None else str(c) for c in row]
            else:
                rows.append(["" if c is None else c for c in row])
            if i >= max_rows:
                truncated = True
                break
        return {
            "sheet": ws.title,
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
            "truncated": truncated,
        }
    finally:
        wb.close()


def write_xlsx(
    path: str | Path,
    rows: list[list[Any]],
    sheet: str = "Sheet1",
) -> str:
    """将二维列表写入 xlsx（首行可作表头），返回落盘绝对路径。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for row in rows:
        ws.append(row)
    dest = Path(path)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return str(dest.resolve())
